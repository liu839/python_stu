import openpyxl
import os
import easygui as eg
dict_={'序号':'A','班级':'B','姓名':'C','语文':'D','数学':'E','外语':'F','生物':'G','政治':'H','历史':'I','地理':'J','化学':'K','物理':'H'}
list_=['A','B','C','D','E','F','G','H','I']
filename1=eg.fileopenbox(msg="选择主文件",title="文件转录")
dirname=eg.diropenbox(msg="选择副文件夹",title="文件转录")
os.chdir(dirname)

wb1=openpyxl.load_workbook(filename1)
ws1=wb1["年级"]

list_dirname=os.listdir(dirname)

for eachname in list_dirname:
    wb2=openpyxl.load_workbook(eachname)
    ws2=wb2["学生答题详情表"]
    print("正在判断%s"%(eachname))
    for each in ws1['A1':'L1']:
        #识别目前在判断什么成绩
        for i in each:
            if i.value in eachname:
                score_1_col=dict_[i.value]
                score_2_col=score_1_col
                break
        score_2_col='D'
    for i_1 in range(2,ws1.max_row+1):
    #实际转录模块
        name_1_sign='C'+str(i_1)
        name_1=ws1[name_1_sign].value
        score_1_sign=score_1_col+str(i_1)
        score_1=ws1[score_1_sign].value
        if score_1!=None:
            continue
        for i_2 in range(2,ws2.max_row+1):
            name_2_sign='C'+str(i_2)
            name_2=ws2[name_2_sign].value
            score_2_sign=score_2_col+str(i_2)
            score_2=ws2[score_2_sign].value
            if name_1==name_2:
                ws1[score_1_sign].value=ws2[score_2_sign].value
                break


wb1.save(r'C:\Users\71037\Desktop\总表.xlsx')